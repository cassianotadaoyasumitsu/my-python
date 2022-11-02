import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
prod_list = inv_file["Sheet1"]

prod_per_supplier = {}
total_value_per_supplier = {}
prod_under_min = {}
# add value to unique cell
prod_list.cell(row=1, column=5).value = "Total"

for prod_row in range(2, prod_list.max_row + 1):
    supplier_name = prod_list.cell(prod_row, 4).value
    inventory = prod_list.cell(prod_row, 2).value
    price = prod_list.cell(prod_row, 3).value
    prod_num = prod_list.cell(prod_row, 1).value
    inventory_price = prod_list.cell(prod_row, 5)

    # calculation number of products per supplier
    if supplier_name in prod_per_supplier:
        current_num_prod = prod_per_supplier.get(supplier_name)
        prod_per_supplier[supplier_name] = current_num_prod + 1
    else:
        prod_per_supplier[supplier_name] = 1

    # total value per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # products under minimal amount
    if inventory <= 10:
        prod_under_min[int(prod_num)] = int(inventory)

    # add new column total inventory price
    inventory_price.value = inventory * price

inv_file.save("inventory_with_total.xlsx")

print(prod_per_supplier)
print(total_value_per_supplier)
print(prod_under_min)
